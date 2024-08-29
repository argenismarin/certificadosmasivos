from openpyxl import load_workbook
from docx import Document
from docx2pdf import convert  

# Se requiere la biblioteca docx2pdf

# Ruta del archivo Excel con los datos de los participantes
excel_file = r".\cor\cor.xlsx"

# Ruta de la plantilla del certificado
cert_template = r".\cor\Formato certificado.docx"

# Ruta de la carpeta para guardar los certificados en PDF
output_folder = r".\Certificados\pdf"

# Cargar el archivo Excel
wb = load_workbook(excel_file)
sheet = wb.active


# Leer los datos de los participantes desde el archivo Excel
participants = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    name = row[0]
    cedula = row[1]
    participants.append((name, cedula))

# Procesar cada participante y generar el certificado
for participant in participants:
    name, cedula = participant

    # Crear una copia de la plantilla del certificado
    document = Document(cert_template)

    # Reemplazar los marcadores en la plantilla con los datos del participante
    for paragraph in document.paragraphs:
        if "{{name}}" in paragraph.text:
            paragraph.text = paragraph.text.replace("{{name}}", name)
        if "{{cedula}}" in paragraph.text:
            paragraph.text = paragraph.text.replace("{{cedula}}", str(cedula))

    # Guardar el certificado en formato DOCX
    cert_filename = f"{name}_certificado.docx"
    cert_path = f"{output_folder}/{cert_filename}"
    document.save(cert_path)

    # Convertir el certificado a PDF
    pdf_filename = f"{name}_certificado.pdf"
    pdf_path = f"{output_folder}/{pdf_filename}"
    convert(cert_path, pdf_path)

    print(f"Se generó el certificado para {name} (Cédula: {cedula})")

print("Proceso completado.")